"""
Modul untuk generate laporan Excel (.xlsx) berisi data mentah, ringkasan metrik,
dan chart tertanam sebagai gambar. FITUR PREMIUM.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def generate_excel(title: str, period: str, brand_color: str, summary: dict,
                    chart_images: list, df) -> bytes:
    """
    chart_images: list of tuples (chart_title, png_bytes)
    Return: bytes dari file .xlsx
    """
    wb = Workbook()
    header_fill = PatternFill(start_color=brand_color.lstrip("#"), end_color=brand_color.lstrip("#"), fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color=brand_color.lstrip("#"))

    # --- Sheet 1: Ringkasan ---
    ws_summary = wb.active
    ws_summary.title = "Ringkasan"
    ws_summary["A1"] = title
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = period
    ws_summary["A2"].font = Font(italic=True, color="666666")

    row = 4
    if summary:
        headers = ["Metrik", "Total", "Rata-rata", "Min", "Max"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_summary.cell(row=row, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        row += 1
        for name, stats in summary.items():
            ws_summary.cell(row=row, column=1, value=name)
            ws_summary.cell(row=row, column=2, value=round(stats["total"], 2))
            ws_summary.cell(row=row, column=3, value=round(stats["average"], 2))
            ws_summary.cell(row=row, column=4, value=round(stats["min"], 2))
            ws_summary.cell(row=row, column=5, value=round(stats["max"], 2))
            row += 1

    for col_idx in range(1, 6):
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = 20

    # --- Sheet 2: Chart (gambar ditempel) ---
    if chart_images:
        ws_charts = wb.create_sheet("Chart")
        img_row = 1
        for chart_title, png_bytes in chart_images:
            ws_charts.cell(row=img_row, column=1, value=chart_title).font = Font(bold=True, size=12)
            img_row += 1
            img_stream = io.BytesIO(png_bytes)
            xl_img = XLImage(img_stream)
            xl_img.width = 600
            xl_img.height = 300
            ws_charts.add_image(xl_img, f"A{img_row}")
            img_row += 18  # kasih jarak sebelum chart berikutnya

    # --- Sheet 3: Data mentah ---
    ws_data = wb.create_sheet("Data Mentah")
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, (_, data_row) in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(data_row, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=str(value) if not isinstance(value, (int, float)) else value)

    for col_idx in range(1, len(df.columns) + 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = 16

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
